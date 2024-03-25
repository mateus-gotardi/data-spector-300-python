from flask import Flask, request, jsonify
from flask_cors import CORS
from functions import processar_linha, make_yy_mm, calc_data
from openpyxl import load_workbook

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "https://data-spector-300-vue.vercel.app"}})
@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return 'Nenhum arquivo enviado', 400
    file = request.files['file']
    if file.filename.endswith('.xlsx'):
        wb = load_workbook(file)
        ws = wb.active
        file_data = []
        for row in ws.iter_rows(min_row=2):
            linha = ",".join([str(cell.value) for cell in row]) 
            file_data.append(linha)
        file_content = '\n'.join(file_data)
    else:
        file_content = file.read().decode('utf-8')
    linhas = file_content.split('\n')
    cabecalho = linhas[0]
    data_processados = []
    data = linhas[1:]
    for linha in data:
        if file.filename.endswith('.xlsx'):
            colunas = processar_linha(linha, 'xlsx')
        else:
            colunas = processar_linha(linha, 'csv')
        if colunas[1] == '30':
            data_processados.append(colunas)
    data_ordenados = sorted(data_processados, key=lambda x: make_yy_mm(x[2]))
    mrr_data = calc_data(data_ordenados)
    return jsonify(mrr_data), 200

if __name__ == '__main__':
    app.run(debug=True)
